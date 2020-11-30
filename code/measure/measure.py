import os
from openpyxl import load_workbook, Workbook
from saliency_toolbox import calculate_measures

sm_dir = "output/MINet_Res50_S320_BS4_E50_WE1_AMPn_LR0.001_LTpoly_OTsgdtrick_ALy_BIy_MSn"
# dataset = "DUT-OMRON"
gt_dir = "../../MINet-Datasets"

# List of datasets
# ["DUTS", "DUT-OMRON", "HKU-IS", "ECSSD", "PASCAL-S", "SOC"]
dataset_list = ["DUTS", "DUT-OMRON", "HKU-IS", "ECSSD", "PASCAL-S", "SOC"]

# List of metrics
# ['MAE', 'E-measure', 'S-measure', 'Max-F', 'Adp-F', 'Wgt-F']
metric_list = ['Max-F', 'Adp-F', 'Wgt-F','E-measure', 'S-measure','MAE']

print("metric_list:" + str(metric_list))

wb = Workbook()
sheet = wb.create_sheet(title="Mesure Results", index=0)

for i, metric in enumerate(metric_list):
    pos_c = f"{chr(ord('A') + (i + 1) % 26)}"
    pos_r = f"{1}"
    sheet[pos_c+pos_r] = metric


for i, dataset in enumerate(dataset_list):

    print("measuring:" + str(dataset))
    pos_r = f"{i + 2}"
    sheet["A"+pos_r] = str(dataset)

    sm = os.path.join(sm_dir, "pre", dataset.lower())
    
    # DUTS and SOC need special treatment due to its folder structure
    if dataset == "DUTS":
        gt = os.path.join(gt_dir, dataset, "Test", "Mask")
    elif dataset == "SOC":
        gt = os.path.join(gt_dir, dataset, "Validation", "Mask")
    else:
        gt = os.path.join(gt_dir, dataset, "Mask")

    # sm_dir = "output/MINet_VGG16_S320_BS4_E50_WE1_AMPn_LR0.001_LTpoly_OTsgdtrick_ALy_BIy_MSn/pre/dut-omron"
    # gt_dir = "../../MINet-Datasets/DUT-OMRON/Mask"

    res, pr = calculate_measures(gt, sm, metric_list, save=False)
    print(res)

    for j, (key, value) in enumerate(res.items()):
        pos_c = f"{chr(ord('A') + (j + 1) % 26)}"
        sheet[pos_c+pos_r] = str(value)

    # if pr:
    #     print(pr)

wb.save("output/Mesure Results Res50.xlsx")