# @Time    : 2020/7/4
# @Author  : Lart Pang
# @Email   : lartpang@163.com
# @File    : recoder.py
# @Project : utils/recoder.py
# @GitHub  : https://github.com/lartpang
import functools
import os
from datetime import datetime

from openpyxl import load_workbook, Workbook
from torch.utils.tensorboard import SummaryWriter
from torchvision.utils import make_grid

from utils.misc import check_mkdir, construct_print


class TBRecorder(object):
    def __init__(self, tb_path):
        check_mkdir(dir_path=tb_path)

        self.tb = SummaryWriter(tb_path)

    def record_curve(self, name, data, curr_iter):
        if not isinstance(data, (tuple, list)):
            self.tb.add_scalar(f"data/{name}", data, curr_iter)
        else:
            for idx, data_item in enumerate(data):
                self.tb.add_scalar(f"data/{name}_{idx}", data_item[name], curr_iter)

    def record_image(self, name, data, curr_iter):
        data_grid = make_grid(data, nrow=data.size(0), padding=5)
        self.tb.add_image(name, data_grid, curr_iter)

    def close_tb(self):
        self.tb.close()


class XLSXRecoder(object):
    def __init__(self, xlsx_path, module_name, model_name):
        self.dataset_list = ["DUTS", "DUT-OMRON", "HKU-IS", "ECSSD", "PASCAL-S", "SOC","MSRA10K","THUR15K"]
        self.dataset_num_list = [5019, 5168, 1447, 1000, 850, 1200, 10000, 15531]
        # self.metric_list = ["MAXF", "MEANF", "MAE"]
        self.metric_list = ["MAXF", "MEANF", "MAE",'Max-F', 'Adp-F', 'Wgt-F', 'E-measure', 'S-measure', 'MAE2']

        self.module_name = module_name
        self.model_name = model_name

        self.path = xlsx_path
        if not os.path.exists(self.path):
            self.create_xlsx()

    def create_xlsx(self):
        num_metrics = len(self.metric_list)
        num_datasets = len(self.dataset_list)

        # Create a workbook object
        wb = Workbook()
        # Create a worksheet object
        sheet = wb.create_sheet(title="Results", index=0)
        # Add row labels
        sheet["A1"] = "name_dataset"
        sheet["A2"] = "num_dataset"

        for i, dataset_name in enumerate(self.dataset_list):
            if (i * num_metrics + 1) // 26 == 0:
                start_region_idx = f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
            else:
                start_region_idx = (
                    f"{chr(ord('A') + (i * num_metrics + 1) // 26 - 1)}"
                    f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
                )
            if ((i + 1) * num_metrics) // 26 == 0:
                end_region_idx = f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
            else:
                end_region_idx = (
                    f"{chr(ord('A') + ((i + 1) * num_metrics) // 26 - 1)}"
                    f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
                )
            region_idx = f"{start_region_idx}:{end_region_idx}"
            sheet.merge_cells(region_idx)  # merge cells for each dataset heading
            sheet[start_region_idx] = dataset_name.upper()

            # Construct the second row of data
            start_region_idx = start_region_idx.replace("1", "2")
            sheet[start_region_idx] = self.dataset_num_list[i]

        # Construct the thrid row of data
        third_row = ["metrics"] + self.metric_list * num_datasets
        sheet.append(third_row)

        # Create a second worksheet object
        sheet = wb.create_sheet(title=self.module_name, index=0)
        # Add row labels
        sheet["A1"] = self.model_name
        sheet.merge_cells("A2:B2")
        sheet["A2"] = "Datasets"
        sheet["A3"] = "name_dataset"
        sheet["B3"] = "num_dataset"

        for i, dataset_name in enumerate(self.dataset_list):
            sheet[f"A{i+4}"] = dataset_name.upper()
            sheet[f"B{i+4}"] = self.dataset_num_list[i]

        # Construct Measurement headings
        sheet["C2"] = "Metrics"
        for i, metric_name in enumerate(self.metric_list):
            sheet[f"{chr(ord('C') + i)}3"] = metric_name

        # Save the workbook
        wb.save(self.path)

    def write_xlsx(self, model_name, data):
        """
        Write data to xlsx file

        :param model_name: Model name
        :param data: Data information, including the name of the data set and the corresponding test results
        """

        num_metrics = len(self.metric_list)
        num_datasets = len(self.dataset_list)

        # You must first create the xlsx file from the previous function to ensure that the first 
        # three rows meet the requirements, and the subsequent operations all start from the fourth line
        wb = load_workbook(self.path)
        assert "Results" in wb.sheetnames, (
            "Please make sure you are " "working with xlsx files " "created by `create_xlsx`"
        )
        sheet = wb["Results"]
        num_cols = sheet.max_column

        model_in_file = False
        for i in range(sheet.max_row):
            if sheet.cell(row=i+1,column=1).value == model_name:
                model_in_file = True
                idx_insert_row = i+1

        if model_in_file == False:
            # If model name already exists in spreadsheet, only need to update the corresponding data set results
            idx_insert_row = len(sheet["A"]) + 1
            sheet.cell(row=idx_insert_row, column=1, value=model_name)

        for dataset_name in data.keys():
            # Loop through each cell
            dataset_found = False
            for row in sheet.iter_rows(min_row=1, min_col=2, max_col=num_cols, max_row=1):
                for cell in row:
                    if cell.value == dataset_name.upper():
                        dataset_found = True
                        for i in range(num_metrics):
                            metric_name = sheet.cell(row=3, column=cell.column + i).value
                            sheet.cell(
                                row=idx_insert_row,
                                column=cell.column + i,
                                value=data[dataset_name][metric_name],
                            )
            # Data set not in spreadsheet yet
            if dataset_found == False:
                insertcol = sheet.max_column

                if insertcol // 26 == 0:
                    start_region_idx = f"{chr(ord('A') + insertcol)}1"
                else:
                    start_region_idx = (
                        f"{chr(ord('A') + insertcol // 26 - 1)}"
                        f"{chr(ord('A') + insertcol % 26)}1"
                    )
                if (insertcol + num_metrics - 1) // 26 == 0:
                    end_region_idx = f"{chr(ord('A') + insertcol + num_metrics - 1)}1"
                else:
                    end_region_idx = (
                        f"{chr(ord('A') + (insertcol + num_metrics - 1) // 26 - 1)}"
                        f"{chr(ord('A') + (insertcol + num_metrics - 1) % 26)}1"
                    )

                region_idx = f"{start_region_idx}:{end_region_idx}"
                sheet.merge_cells(region_idx)  # merge cells for each dataset heading
                sheet[start_region_idx] = dataset_name.upper()

                # Add number of data items
                if dataset_name.upper() in self.dataset_list:
                    for i, dataset_name2 in enumerate(self.dataset_list):
                        if dataset_name2 == dataset_name.upper():
                            start_region_idx = start_region_idx.replace("1", "2")
                            sheet[start_region_idx] = self.dataset_num_list[i]

                # Add measurement titles
                for i, measure in enumerate(self.metric_list):
                    sheet.cell(row=3,column=i+insertcol + 1,value=measure)

                # Add measurement data
                for row in sheet.iter_rows(min_row=idx_insert_row, min_col=insertcol + 1, max_col=insertcol + num_metrics, max_row=idx_insert_row):
                    for cell in row:
                        metric_name = sheet.cell(row=3, column=cell.column).value
                        cell.value = data[dataset_name][metric_name]

        # write to second worksheet
        if self.module_name in wb.sheetnames:
            sheet = wb[self.module_name]
        else:
            sheet = wb.create_sheet(self.module_name, index=0)

            # Add row labels
            sheet["A1"] = self.model_name
            sheet.merge_cells("A2:B2")
            sheet["A2"] = "Datasets"
            sheet["A3"] = "name_dataset"
            sheet["B3"] = "num_dataset"

            for i, dataset_name in enumerate(self.dataset_list):
                sheet[f"A{i+4}"] = dataset_name.upper()
                sheet[f"B{i+4}"] = self.dataset_num_list[i]

            # Construct Measurement headings
            sheet["C2"] = "Metrics"
            for i, metric_name in enumerate(self.metric_list):
                sheet[f"{chr(ord('C') + i)}3"] = metric_name

        for dataset_name in data.keys():
            dataset_found = False
            for row in sheet.iter_rows(min_row=4, min_col=3, max_col=num_metrics+2, max_row=sheet.max_row):
                for cell in row:
                    if sheet.cell(row=cell.row, column=1).value == dataset_name.upper():
                        dataset_found = True
                        metric_name = sheet.cell(row=3, column=cell.column).value
                        sheet.cell(
                            row=cell.row,
                            column=cell.column,
                            value=data[dataset_name][metric_name],
                        )

            if dataset_found == False:
                # Add new dataset to spreadsheet
                insertrow = sheet.max_row + 1
                sheet.cell(row=insertrow, column=1, value=dataset_name.upper())

                # Add number of data items
                if dataset_name.upper() in self.dataset_list:
                    for i, dataset_name2 in enumerate(self.dataset_list):
                        if dataset_name2 == dataset_name.upper():
                            sheet.cell(row=insertrow, column=2, value=self.dataset_num_list[i])

                for row in sheet.iter_rows(min_row=insertrow, min_col=3, max_col=num_metrics+2, max_row=insertrow):
                    for cell in row:
                        metric_name = sheet.cell(row=3, column=cell.column).value
                        sheet.cell(
                            row=cell.row,
                            column=cell.column,
                            value=data[dataset_name][metric_name],
                        )

        wb.save(self.path)


def Timer(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = datetime.now()
        construct_print(f"a new epoch start: {start_time}")
        func(*args, **kwargs)
        construct_print(f"the time of the epoch: {datetime.now() - start_time}")

    return wrapper
